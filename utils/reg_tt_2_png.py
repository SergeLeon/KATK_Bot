from pathlib import Path
from pprint import pprint

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.styles import DEFAULT_FONT
from openpyxl.styles import Alignment

import excel2img

import sys

# Для возможности вызова напрямую
sys.path.append(str(Path(__file__).parent.parent.absolute()))

from config import REGULAR_TIMETABLE_PATH
from parsers.xlsx_parser import _worksheet_as_list, _extract_timetables_by_weekdays

HEADER_ADDITION = "с 13 января 2023 г. по 30 июня 2023 г.(изменения от февраля 2023)"

OUTPUT_FOLDER = "utils/tables"
OUTPUT_FOLDER = Path(OUTPUT_FOLDER)
OUTPUT_FOLDER.mkdir(exist_ok=True)

font = Font(
    name="Calibri",
    size=9,
    bold=False,
    italic=False,
    vertAlign=None,
    underline="none",
    strike=False
)

for k, v in font.__dict__.items():
    setattr(DEFAULT_FONT, k, v)


def clear_list_end(str_list: list):
    str_list = list(str_list)
    for i in str_list[::-1]:
        if i:
            break
        str_list.pop(-1)
    return str_list


def auto_width(sheet):
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter

        for cell in col:
            if cell.coordinate in sheet.merged_cells:
                continue
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.02
        sheet.column_dimensions[column].width = adjusted_width
    return sheet


def get_tables(filename):
    wb = load_workbook(filename=filename, read_only=True)
    ws = wb.worksheets[0]

    table = _worksheet_as_list(ws)
    output_tables = {}

    timetables = _extract_timetables_by_weekdays(table)
    for weekday, tables in timetables.items():
        for group, table in tables.items():
            if group not in output_tables:
                output_tables[group] = []
            table = list(table)
            table[1] = clear_list_end(table[1])
            output_tables[group].append((weekday, *table[1]))
    return output_tables


def main():
    tables = get_tables(REGULAR_TIMETABLE_PATH)
    pprint(tables)

    wb = Workbook()

    for group_name, table in tables.items():

        sheet = wb.create_sheet(group_name)
        pars = len(max(table, key=len))
        for i in range(pars - 1):
            sheet.cell(i + 3, 1, str(i + 1))

        for column, row in enumerate(table, 2):
            for line, cell in enumerate(row, 2):
                sheet.cell(line, column, cell.title())

        sheet = auto_width(sheet)

        cell_width = sum(sheet.column_dimensions[letter].width for letter in "ABCDEFG") * 1.1

        header = f"{HEADER_ADDITION} | {group_name.rjust(int(cell_width / 2 - len(HEADER_ADDITION)), '_')}"

        cell = sheet.cell(1, 1, header)
        cell.alignment = Alignment(horizontal='left')

        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)

    wb.remove_sheet(wb.get_sheet_by_name('Sheet'))

    timetable_by_group_xlsx = Path(__file__).resolve().parent / "tablex.xlsx"
    print(timetable_by_group_xlsx)

    wb.save(timetable_by_group_xlsx)

    for group_name in tables:
        excel2img.export_img(timetable_by_group_xlsx, f"{OUTPUT_FOLDER / group_name}.png", group_name, None)
        print(f"{OUTPUT_FOLDER / group_name}.png")


if __name__ == "__main__":
    main()
